"""
Import foods from Excel file into the database.
Avoids duplicates by checking existing food names.
"""
import sys
import openpyxl
from sqlalchemy import create_engine, text
from sqlalchemy.orm import sessionmaker

DB_URL = "sqlite:///./sukarak_v3.db"
EXCEL_PATH = r"E:\dr.moahmed\food.xlsx"

def import_foods():
    engine = create_engine(DB_URL)
    Session = sessionmaker(bind=engine)
    db = Session()

    # Ensure table exists
    db.execute(text("""
        CREATE TABLE IF NOT EXISTS food_db (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name VARCHAR(255) NOT NULL,
            type VARCHAR(255),
            serving VARCHAR(255),
            glycemic_index VARCHAR(50),
            calories VARCHAR(50),
            carb VARCHAR(50),
            protein VARCHAR(50),
            fats VARCHAR(50)
        )
    """))
    db.commit()

    # Get existing food names to avoid duplicates
    existing = set()
    rows = db.execute(text("SELECT name FROM food_db")).fetchall()
    for row in rows:
        existing.add(row[0].strip() if row[0] else '')
    print(f"Existing foods in DB: {len(existing)}")

    # Read Excel
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active
    headers = [ws.cell(1, i).value for i in range(1, ws.max_column + 1)]
    print(f"Excel headers: {headers}")
    print(f"Excel rows: {ws.max_row - 1}")

    # Map headers to columns
    col_map = {}
    for i, h in enumerate(headers):
        if h:
            h_lower = h.strip().lower()
            if h_lower == 'id':
                col_map['id'] = i + 1
            elif h_lower == 'name':
                col_map['name'] = i + 1
            elif h_lower == 'type':
                col_map['type'] = i + 1
            elif h_lower == 'serving':
                col_map['serving'] = i + 1
            elif 'glycemic' in h_lower or 'gi' in h_lower:
                col_map['glycemic_index'] = i + 1
            elif 'carb' in h_lower:
                col_map['carb'] = i + 1
            elif 'protien' in h_lower or 'protein' in h_lower:
                col_map['protein'] = i + 1
            elif 'fat' in h_lower:
                col_map['fats'] = i + 1
            elif 'calor' in h_lower:
                col_map['calories'] = i + 1

    print(f"Column mapping: {col_map}")

    inserted = 0
    skipped = 0

    for r in range(2, ws.max_row + 1):
        name = ws.cell(r, col_map.get('name', 2)).value
        if not name or not str(name).strip():
            continue

        name = str(name).strip()

        # Skip duplicates
        if name in existing:
            skipped += 1
            continue

        food_type = ws.cell(r, col_map.get('type', 3)).value
        serving = ws.cell(r, col_map.get('serving', 4)).value
        gi = ws.cell(r, col_map.get('glycemic_index', 5)).value
        carb = ws.cell(r, col_map.get('carb', 6)).value
        protein = ws.cell(r, col_map.get('protein', 7)).value
        fats = ws.cell(r, col_map.get('fats', 8)).value
        calories = ws.cell(r, col_map.get('calories', 9)).value

        # Clean values
        food_type = str(food_type).strip() if food_type and str(food_type) != 'None' else None
        serving = str(serving).strip() if serving and str(serving) != 'None' else None
        gi = str(gi).strip() if gi and str(gi) not in ('None', 'NA', 'N/A') else None
        carb = str(carb).strip() if carb and str(carb) not in ('None', 'NA', 'N/A') else None
        protein = str(protein).strip() if protein and str(protein) not in ('None', 'NA', 'N/A') else None
        fats = str(fats).strip() if fats and str(fats) not in ('None', 'NA', 'N/A') else None
        calories = str(calories).strip() if calories and str(calories) not in ('None', 'NA', 'N/A') else None

        db.execute(text("""
            INSERT INTO food_db (name, type, serving, glycemic_index, carb, protein, fats, calories)
            VALUES (:name, :type, :serving, :gi, :carb, :protein, :fats, :calories)
        """), {
            'name': name, 'type': food_type, 'serving': serving,
            'gi': gi, 'carb': carb, 'protein': protein,
            'fats': fats, 'calories': calories
        })
        existing.add(name)
        inserted += 1

    db.commit()
    db.close()

    # Final count
    db2 = Session()
    total = db2.execute(text("SELECT COUNT(*) FROM food_db")).scalar()
    cats = db2.execute(text("SELECT type, COUNT(*) FROM food_db GROUP BY type ORDER BY COUNT(*) DESC")).fetchall()
    db2.close()

    print(f"\n{'='*50}")
    print(f"✅ Import complete!")
    print(f"   Inserted: {inserted}")
    print(f"   Skipped (duplicates): {skipped}")
    print(f"   Total in DB: {total}")
    print(f"\n   Categories:")
    for cat, count in cats:
        print(f"     {cat or 'بدون تصنيف'}: {count}")

if __name__ == "__main__":
    import_foods()
